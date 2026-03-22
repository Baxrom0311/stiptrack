from __future__ import annotations

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models.enums import ApplicationStatus, ScholarshipStatus
from app.schemas.admin import ScholarshipResultRow, ScholarshipResultsOut


STATUS_LABELS = {
    ApplicationStatus.DRAFT: "Draft",
    ApplicationStatus.SUBMITTED: "Submitted",
    ApplicationStatus.IN_REVIEW: "In Review",
    ApplicationStatus.WINNER: "Winner",
    ApplicationStatus.REJECTED: "Rejected",
}

SCHOLARSHIP_STATUS_LABELS = {
    ScholarshipStatus.DRAFT: "Draft",
    ScholarshipStatus.OPEN: "Open",
    ScholarshipStatus.CLOSED: "Closed",
    ScholarshipStatus.DONE: "Done",
}


def slugify_filename(value: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", value).strip("-").lower()
    return slug or "scholarship-results"


def _status_label(status: ApplicationStatus) -> str:
    return STATUS_LABELS.get(status, status.value)


def _scholarship_status_label(status: ScholarshipStatus) -> str:
    return SCHOLARSHIP_STATUS_LABELS.get(status, status.value)


def _format_score(value: float | None) -> str:
    if value is None:
        return "-"
    return f"{value:.2f}"


def _format_submitted_at(value) -> str:
    if value is None:
        return "-"
    return value.strftime("%Y-%m-%d %H:%M")


def _row_to_values(row: ScholarshipResultRow) -> list[str]:
    return [
        str(row.rank or "-"),
        row.student_name,
        str(row.student_id),
        str(row.application_id),
        _status_label(row.status),
        _format_score(row.total_score),
        "Yes" if row.is_winner else "No",
        _format_submitted_at(row.submitted_at),
    ]


def build_scholarship_results_excel(results: ScholarshipResultsOut) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Results"

    worksheet.append(["Scholarship", results.scholarship_title])
    worksheet.append(["Status", _scholarship_status_label(results.scholarship_status)])
    worksheet.append(["Winner slots", results.max_winners])
    worksheet.append(["Current winners", results.winners_count])
    worksheet.append([])

    header_row_index = 6
    headers = ["Rank", "Student", "Student ID", "Application ID", "Status", "Total Score", "Winner", "Submitted At"]
    worksheet.append(headers)
    for row in results.rows:
        worksheet.append(_row_to_values(row))

    header_fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    header_font = Font(bold=True)
    for cell in worksheet[header_row_index]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = "A7"

    column_widths: dict[str, int] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            column_widths[cell.column_letter] = min(max(column_widths.get(cell.column_letter, 0), length + 2), 40)

    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_scholarship_results_pdf(results: ScholarshipResultsOut) -> bytes:
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()

    story = [
        Paragraph("Scholarship Results Export", styles["Title"]),
        Spacer(1, 8),
        Paragraph(f"Scholarship: {results.scholarship_title}", styles["Normal"]),
        Paragraph(f"Status: {_scholarship_status_label(results.scholarship_status)}", styles["Normal"]),
        Paragraph(f"Winner slots: {results.max_winners}", styles["Normal"]),
        Paragraph(f"Current winners: {results.winners_count}", styles["Normal"]),
        Spacer(1, 10),
    ]

    table_data = [
        ["Rank", "Student", "Student ID", "Application ID", "Status", "Score", "Winner", "Submitted At"],
        *[_row_to_values(row) for row in results.rows],
    ]
    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[14 * mm, 46 * mm, 34 * mm, 46 * mm, 26 * mm, 20 * mm, 18 * mm, 32 * mm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2E8F0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ]
        )
    )
    story.append(table)

    document.build(story)
    return buffer.getvalue()
